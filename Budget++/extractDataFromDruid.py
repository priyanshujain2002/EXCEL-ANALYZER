import requests
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Disable telemetry before importing CrewAI to prevent connection issues
os.environ["OTEL_SDK_DISABLED"] = "true"
os.environ["CREWAI_TELEMETRY_ENABLED"] = "false"
os.environ["CREWAI_DISABLE_TELEMETRY"] = "true"

from crewai import Agent, Task, Crew, LLM
from crewai.knowledge.source.excel_knowledge_source import ExcelKnowledgeSource
import boto3
import logging
from typing import List, Dict, Any

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Bedrock session and LLM configuration
session = boto3.Session(region_name="us-east-1")

llm = LLM(
    model="bedrock/us.anthropic.claude-3-7-sonnet-20250219-v1:0",
    aws_region_name="us-east-1",
    temperature=0.3,  # Moderate temperature for creative but focused descriptions
    timeout=60,
    max_retries=2,
)

# Embedder configuration for knowledge source
embedder_config = {
    "provider": "bedrock",
    "config": {
        "model": "amazon.titan-embed-text-v1",
        "session": session
    }
}

# State abbreviation to full name mapping (case-insensitive)
state_mapping = {
    "AA": "Armed Forces Americas",
    "AE": "Armed Forces Europe, Middle East, & Canada",
    "AK": "Alaska",
    "AL": "Alabama",
    "AP": "Armed Forces Pacific",
    "AR": "Arkansas",
    "AS": "American Samoa",
    "AZ": "Arizona",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DC": "District of Columbia",
    "DE": "Delaware",
    "FL": "Florida",
    "FM": "Federated States of Micronesia",
    "GA": "Georgia",
    "GU": "Guam",
    "HI": "Hawaii",
    "IA": "Iowa",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "MA": "Massachusetts",
    "MD": "Maryland",
    "ME": "Maine",
    "MH": "Marshall Islands",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MO": "Missouri",
    "MP": "Northern Mariana Islands",
    "MS": "Mississippi",
    "MT": "Montana",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "NE": "Nebraska",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NV": "Nevada",
    "NY": "New York",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "PW": "Palau",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VA": "Virginia",
    "VI": "Virgin Islands",
    "VT": "Vermont",
    "WA": "Washington",
    "WI": "Wisconsin",
    "WV": "West Virginia",
    "WY": "Wyoming"
}

url = "https://druid.use1-rprod.k8s.adgear.com/druid/v2/sql"

payload = json.dumps({
    "query": """
WITH split_response AS (
    SELECT 
        LOWER(TRIM(genre_individual)) as genre,
        LOWER(TRIM(
            CASE 
                WHEN UPPER(TRIM(genre_individual)) = 'AP' THEN 'Audience Participation'
                WHEN UPPER(TRIM(genre_individual)) = 'AC' THEN 'Award Ceremonies & Pageants'
                WHEN UPPER(TRIM(genre_individual)) = 'CP' THEN 'Children''s Programming'
                WHEN UPPER(TRIM(genre_individual)) = 'CV' THEN 'Comedy Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'CM' THEN 'Concert Music'
                WHEN UPPER(TRIM(genre_individual)) = 'CC' THEN 'Conversation & Colloquies'
                WHEN UPPER(TRIM(genre_individual)) = 'DD' THEN 'Daytime Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'D' THEN 'Devotional'
                WHEN UPPER(TRIM(genre_individual)) = 'DO' THEN 'Documentary & General'
                WHEN UPPER(TRIM(genre_individual)) = 'DN' THEN 'Documentary & News'
                WHEN UPPER(TRIM(genre_individual)) = 'EA' THEN 'Evening Animation'
                WHEN UPPER(TRIM(genre_individual)) = 'FF' THEN 'Feature Film'
                WHEN UPPER(TRIM(genre_individual)) = 'GD' THEN 'General Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'GV' THEN 'General Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'IA' THEN 'Instructions & Advice'
                WHEN UPPER(TRIM(genre_individual)) = 'MD' THEN 'Musical Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'N' THEN 'News'
                WHEN UPPER(TRIM(genre_individual)) = 'OP' THEN 'Official Police'
                WHEN UPPER(TRIM(genre_individual)) = 'P' THEN 'Paid Political'
                WHEN UPPER(TRIM(genre_individual)) = 'PV' THEN 'Participation Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'PC' THEN 'Popular Music'
                WHEN UPPER(TRIM(genre_individual)) = 'PD' THEN 'Private Detective'
                WHEN UPPER(TRIM(genre_individual)) = 'QG' THEN 'Quiz -Give Away'
                WHEN UPPER(TRIM(genre_individual)) = 'QP' THEN 'Quiz -Panel'
                WHEN UPPER(TRIM(genre_individual)) = 'SF' THEN 'Science Fiction'
                WHEN UPPER(TRIM(genre_individual)) = 'CS' THEN 'Situation Comedy'
                WHEN UPPER(TRIM(genre_individual)) = 'SA' THEN 'Sports Anthology'
                WHEN UPPER(TRIM(genre_individual)) = 'SC' THEN 'Sports Commentary'
                WHEN UPPER(TRIM(genre_individual)) = 'SE' THEN 'Sports Event'
                WHEN UPPER(TRIM(genre_individual)) = 'SN' THEN 'Sports News'
                WHEN UPPER(TRIM(genre_individual)) = 'SM' THEN 'Suspense/Mystery'
                WHEN UPPER(TRIM(genre_individual)) = 'EW' THEN 'Western Drama'
                ELSE TRIM(genre_individual)
            END
        )) as genre_updated,
        creativetype,
        de_region,
        de_country,
        "row_count"
    FROM (
        SELECT 
            CASE WHEN genre IS NULL THEN 'Unknown' ELSE genre END as genre,
            creativetype,
            de_region,
            de_country,
            "row_count"
        FROM "ctv_untargeted_bid_response"
        WHERE 
            de_country = 'US'
            AND exchange_id = 55
            AND __time >= '2025-09-03T00:00:00.000Z'
            AND __time <= '2025-09-09T00:00:00.000Z'
    ) CROSS JOIN UNNEST(STRING_TO_ARRAY(genre, ',')) AS t(genre_individual)
    WHERE UPPER(TRIM(genre_individual)) != 'A'
),
split_request AS (
    SELECT 
        LOWER(TRIM(genre_individual)) as genre,
        LOWER(TRIM(
            CASE 
                WHEN UPPER(TRIM(genre_individual)) = 'AP' THEN 'Audience Participation'
                WHEN UPPER(TRIM(genre_individual)) = 'AC' THEN 'Award Ceremonies & Pageants'
                WHEN UPPER(TRIM(genre_individual)) = 'CP' THEN 'Children''s Programming'
                WHEN UPPER(TRIM(genre_individual)) = 'CV' THEN 'Comedy Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'CM' THEN 'Concert Music'
                WHEN UPPER(TRIM(genre_individual)) = 'CC' THEN 'Conversation & Colloquies'
                WHEN UPPER(TRIM(genre_individual)) = 'DD' THEN 'Daytime Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'D' THEN 'Devotional'
                WHEN UPPER(TRIM(genre_individual)) = 'DO' THEN 'Documentary & General'
                WHEN UPPER(TRIM(genre_individual)) = 'DN' THEN 'Documentary & News'
                WHEN UPPER(TRIM(genre_individual)) = 'EA' THEN 'Evening Animation'
                WHEN UPPER(TRIM(genre_individual)) = 'FF' THEN 'Feature Film'
                WHEN UPPER(TRIM(genre_individual)) = 'GD' THEN 'General Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'GV' THEN 'General Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'IA' THEN 'Instructions & Advice'
                WHEN UPPER(TRIM(genre_individual)) = 'MD' THEN 'Musical Drama'
                WHEN UPPER(TRIM(genre_individual)) = 'N' THEN 'News'
                WHEN UPPER(TRIM(genre_individual)) = 'OP' THEN 'Official Police'
                WHEN UPPER(TRIM(genre_individual)) = 'P' THEN 'Paid Political'
                WHEN UPPER(TRIM(genre_individual)) = 'PV' THEN 'Participation Variety'
                WHEN UPPER(TRIM(genre_individual)) = 'PC' THEN 'Popular Music'
                WHEN UPPER(TRIM(genre_individual)) = 'PD' THEN 'Private Detective'
                WHEN UPPER(TRIM(genre_individual)) = 'QG' THEN 'Quiz -Give Away'
                WHEN UPPER(TRIM(genre_individual)) = 'QP' THEN 'Quiz -Panel'
                WHEN UPPER(TRIM(genre_individual)) = 'SF' THEN 'Science Fiction'
                WHEN UPPER(TRIM(genre_individual)) = 'CS' THEN 'Situation Comedy'
                WHEN UPPER(TRIM(genre_individual)) = 'SA' THEN 'Sports Anthology'
                WHEN UPPER(TRIM(genre_individual)) = 'SC' THEN 'Sports Commentary'
                WHEN UPPER(TRIM(genre_individual)) = 'SE' THEN 'Sports Event'
                WHEN UPPER(TRIM(genre_individual)) = 'SN' THEN 'Sports News'
                WHEN UPPER(TRIM(genre_individual)) = 'SM' THEN 'Suspense/Mystery'
                WHEN UPPER(TRIM(genre_individual)) = 'EW' THEN 'Western Drama'
                ELSE TRIM(genre_individual)
            END
        )) as genre_updated,
        creativetype,
        de_region,
        de_country,
        "row_count"
    FROM (
        SELECT 
            CASE WHEN genre IS NULL THEN 'Unknown' ELSE genre END as genre,
            creativetype,
            de_region,
            de_country,
            "row_count"
        FROM "ctv_untargeted_bid_request"
        WHERE 
            exchange_id = 55
            AND de_country = 'US'
            AND __time >= '2025-09-03T00:00:00.000Z'
            AND __time <= '2025-09-09T00:00:00.000Z'
    ) CROSS JOIN UNNEST(STRING_TO_ARRAY(genre, ',')) AS t(genre_individual)
    WHERE UPPER(TRIM(genre_individual)) != 'A'
),
cte1 AS (
    SELECT 
        genre,
        genre_updated,
        creativetype,
        de_region,
        de_country,
        SUM("row_count") AS sum_response
    FROM split_response
    GROUP BY 1, 2, 3, 4, 5
),
cte2 AS (
    SELECT
        genre,
        genre_updated,
        creativetype,
        de_region,
        de_country,
        SUM("row_count") AS sum_request
    FROM split_request
    GROUP BY 1, 2, 3, 4, 5
)
SELECT 
    cte1.genre,
    cte1.genre_updated,
    cte1.creativetype,
    cte1.de_region,
    cte1.de_country,
    cte2.sum_request,
    cte1.sum_response
FROM cte1
INNER JOIN cte2 
ON cte1.genre = cte2.genre 
AND cte1.creativetype = cte2.creativetype
AND cte1.de_region = cte2.de_region
AND cte1.de_country = cte2.de_country
"""
})
headers = {
  'Content-Type': 'application/json'
}

print("🔍 Extracting data from Druid...")
response = requests.request("POST", url, headers=headers, data=payload, verify=False)

# Parse JSON response and create dataframe
data = json.loads(response.text)
df = pd.DataFrame(data)

# Debug: Print the columns in the dataframe
print("🔍 Dataframe columns:", df.columns.tolist())

# If there's an error, print the error details
if 'error' in df.columns:
    print("❌ SQL Error detected:")
    print("Error:", df.iloc[0]['error'])
    print("Error Code:", df.iloc[0]['errorCode'])
    print("Error Message:", df.iloc[0]['errorMessage'])
    exit(1)

# Map state abbreviations to full names (case-insensitive)
def map_state_name(region_code):
    if pd.isna(region_code) or region_code == 'Unknown':
        return 'Unknown'
    # Convert to uppercase for case-insensitive matching
    region_code_upper = str(region_code).upper()
    return state_mapping.get(region_code_upper, region_code)  # Return original if not found

# Create de_region_updated column with state mapping applied
# Keep the original de_region column unchanged
df['de_region_updated'] = df['de_region'].apply(map_state_name)

# Add ID column at the beginning (starting from 1)
df.insert(0, 'id', range(1, len(df) + 1))

print("✅ Dataframe created successfully:")
print(df)

# Extract unique genres directly from dataframe (using genre_updated column)
unique_genres = df['genre_updated'].dropna().unique().tolist()
print(f"📋 Found {len(unique_genres)} unique genres")

# Create the genre description agent
genre_description_agent = Agent(
    role='Expert Genre Description Specialist',
    goal='Generate concise, impactful one-line keyword descriptions for entertainment genres that capture their essence and appeal',
    backstory='''You are a master entertainment genre analyst with decades of experience in content categorization,
    audience psychology, and media marketing. Your expertise includes:

    CORE COMPETENCIES:
    - Genre Analysis: Deep understanding of what makes each genre unique and appealing
    - Audience Psychology: Knowledge of what draws viewers to different types of content
    - Marketing Copy: Ability to create compelling, concise descriptions that resonate
    - Keyword Optimization: Skill in identifying the most impactful terms for each genre
    - Trend Awareness: Understanding of current genre trends and audience preferences

    DESCRIPTION PHILOSOPHY:
    - CONCISE: Each description must be exactly one line (under 15 words)
    - KEYWORD-FOCUSED: Use powerful, searchable terms that capture the genre's essence
    - EMOTIONAL: Tap into the core emotional appeal of each genre
    - DISTINCTIVE: Highlight what makes each genre unique from others
    - ACTION-ORIENTED: Use dynamic language that conveys the genre's energy

    DESCRIPTION FRAMEWORK:
    1. Core Theme: What is the central subject matter or premise?
    2. Emotional Tone: What feelings does the genre typically evoke?
    3. Style Elements: What are the characteristic production or narrative styles?
    4. Target Appeal: What type of viewer is most drawn to this genre?

    EXAMPLE DESCRIPTIONS:
    - "Action": High-energy thrills with explosive stunts and heroic adventures
    - "Drama": Emotional character studies exploring complex human relationships
    - "Comedy": Lighthearted humor designed to entertain and amuse audiences
    - "Horror": Suspenseful tales designed to frighten and thrill viewers
    - "Documentary": Factual explorations of real people, places, and events

    You excel at distilling complex genre characteristics into powerful, memorable one-liners.''',
    llm=llm,
    verbose=True,
    allow_delegation=False
)

def generate_genre_descriptions(genre_list: List[str]) -> Dict[str, str]:
    """
    Generate one-line descriptions for the provided genres
    
    Args:
        genre_list: List of genre names
        
    Returns:
        Dictionary mapping genre names to descriptions
    """
    print(f"🎯 Generating descriptions for {len(genre_list)} genres...")
    
    # Process genres in smaller batches to avoid timeout
    batch_size = 50
    all_descriptions = {}
    
    for i in range(0, len(genre_list), batch_size):
        batch_genres = genre_list[i:i + batch_size]
        print(f"🔄 Processing batch {i//batch_size + 1}/{(len(genre_list) + batch_size - 1)//batch_size} ({len(batch_genres)} genres)")
        
        # Create a new task for each batch
        batch_task = Task(
            description=f"""
            Generate compelling one-line keyword descriptions for the following entertainment genres:
            
            GENRES TO PROCESS: {', '.join(batch_genres)}
            
            REQUIREMENTS:
            1. ONE-LINE ONLY: Each description must be exactly one line (under 15 words)
            2. KEYWORD-RICH: Use powerful, searchable terms that capture the genre's essence
            3. EMOTIONAL APPEAL: Tap into the core emotional draw of each genre
            4. DISTINCTIVE: Highlight what makes each genre unique from others
            5. CONSISTENT: Maintain a similar style and tone across all descriptions
            
            PROCESS:
            For each genre in the list:
            1. Analyze the genre's core characteristics and audience appeal
            2. Identify the most impactful keywords and emotional triggers
            3. Craft a concise, powerful one-line description
            4. Ensure the description is unique and distinctive
            
            OUTPUT FORMAT:
            Return a JSON object where keys are the genre names (exactly as provided) and values are the one-line descriptions.
            Example:
            {{
              "general variety": "Diverse entertainment mix with broad audience appeal",
              "drama": "Emotional character studies exploring complex human relationships",
              "news": "Current events and factual reporting with real-time relevance"
            }}
            
            CRITICAL: 
            - Every genre in the provided list must have exactly one description line
            - Use genre names exactly as they appear in the list
            """,
            agent=genre_description_agent,
            expected_output="JSON object with genre names as keys and one-line descriptions as values"
        )
        
        # Create a new crew for each batch
        batch_crew = Crew(
            agents=[genre_description_agent],
            tasks=[batch_task],
            verbose=True
        )
        
        try:
            # Run the crew for this batch
            result = batch_crew.kickoff()
            
            # Parse the JSON result
            try:
                # Handle CrewOutput object by converting to string first
                result_str = str(result)
                
                # Strip markdown code blocks if present
                if result_str.strip().startswith('```json'):
                    # Find the start and end of the JSON content
                    start_idx = result_str.find('```json') + 7  # Skip '```json'
                    end_idx = result_str.rfind('```')
                    if end_idx > start_idx:
                        result_str = result_str[start_idx:end_idx].strip()
                elif result_str.strip().startswith('```'):
                    # Handle generic code blocks
                    start_idx = result_str.find('```') + 3
                    end_idx = result_str.rfind('```')
                    if end_idx > start_idx:
                        result_str = result_str[start_idx:end_idx].strip()
                
                batch_descriptions = json.loads(result_str)
                all_descriptions.update(batch_descriptions)
                logger.info(f"Successfully generated {len(batch_descriptions)} descriptions in batch {i//batch_size + 1}")
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing JSON result in batch {i//batch_size + 1}: {e}")
                logger.error(f"Raw result: {result}")
                continue
                
        except Exception as e:
            logger.error(f"Error generating genre descriptions in batch {i//batch_size + 1}: {e}")
            continue
    
    logger.info(f"Successfully generated {len(all_descriptions)} genre descriptions in total")
    return all_descriptions

# Generate descriptions using CrewAI agent
descriptions = generate_genre_descriptions(unique_genres)

if not descriptions:
    print("❌ Failed to generate genre descriptions")
    exit(1)

print(f"✅ Successfully generated {len(descriptions)} genre descriptions")

# Add description column to dataframe beside genre column
# Find the position of genre column
genre_col_pos = df.columns.get_loc('genre')
# Insert description column right after genre column
df.insert(genre_col_pos + 1, 'description', df['genre_updated'].map(descriptions))
# Fill any NaN values with empty string
df['description'] = df['description'].fillna('')

print("📝 Added description column to dataframe")
print(f"📊 Total rows in dataframe: {len(df)}")
print(f"🔍 Verifying genre descriptions are applied to all rows with same genre:")
# Check a few sample genres to ensure descriptions are applied correctly
sample_genres = df['genre'].dropna().unique()[:3]  # Check first 3 genres
for genre in sample_genres:
    genre_rows = df[df['genre'] == genre]
    unique_descriptions = genre_rows['description'].unique()
    print(f"   Genre '{genre}': {len(genre_rows)} rows, {len(unique_descriptions)} unique description(s)")
    if len(unique_descriptions) == 1:
        print(f"   ✅ All rows have the same description: '{unique_descriptions[0]}'")
    else:
        print(f"   ❌ Multiple descriptions found: {unique_descriptions}")

# Create Excel file with formulas and sorting
excel_filename = "knowledge/druid_query_results_with_descriptions.xlsx"

# Calculate X values for sorting (temporary calculation)
df['unsold_supply_calc'] = df['sum_request'] - df['sum_response']
total_unsold = df['unsold_supply_calc'].sum()
df['x_calc'] = (df['unsold_supply_calc'] / total_unsold * 100) if total_unsold != 0 else 0

# Sort by X values (descending)
df_sorted = df.sort_values('x_calc', ascending=False).reset_index(drop=True)

# Create workbook with sorted data and Excel formulas
wb = Workbook()
ws = wb.active

# Add headers (now including original and updated columns for genre, de_region, and de_country)
headers = ['id', 'genre', 'genre_updated', 'description', 'creativetype', 'de_region', 'de_region_updated', 'de_country', 'sum_request', 'sum_response', 'unsold_supply', 'x', 'y']
ws.append(headers)

# Add sorted data rows
for i, (_, row) in enumerate(df_sorted.iterrows(), 2):
    ws.append([i-1, row['genre'], row['genre_updated'], row['description'], row['creativetype'], row['de_region'], row['de_region_updated'], row['de_country'], row['sum_request'], row['sum_response'], '', '', ''])

# Add Excel formulas to new columns
total_rows = len(df_sorted) + 1
for row in range(2, total_rows + 1):
    # Update ID to sequential
    ws[f'A{row}'] = row - 1
    
    # Column K: Unsold Supply = sum_request - sum_response (now column K because we added genre_updated, de_region_updated, and de_country)
    ws[f'K{row}'] = f'=I{row}-J{row}'
    
    # Column L: X = (Unsold supply of that type/Sum of all unsold supply)*100
    ws[f'L{row}'] = f'=IF(SUM(K$2:K${total_rows})=0,0,(K{row}/SUM(K$2:K${total_rows}))*100)'
    
    # Column M: Y = (Unsold supply of that type / sum_request of that type)*100
    ws[f'M{row}'] = f'=IF(I{row}=0,0,(K{row}/I{row})*100)'

# Ensure the knowledge directory exists
os.makedirs("Budget++/knowledge", exist_ok=True)

# Save Excel file
wb.save(excel_filename)
print(f"🎉 Data saved to Excel file: {excel_filename}")
print("📊 File is sorted by X values (highest to lowest)")
print("📈 Columns added: Description, Unsold Supply, X, Y with Excel formulas (full precision)")

# Show sample descriptions
print("\n📝 Sample descriptions:")
sample_count = min(5, len(descriptions))
for i, (genre, desc) in enumerate(list(descriptions.items())[:sample_count]):
    print(f"   {i+1}. {genre}: {desc}")
