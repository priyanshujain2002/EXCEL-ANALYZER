#!/usr/bin/env python3
"""
Combined Excel Processor - Merges functionality from process_excel.py and multi_sheet_excel_processor.py
Processes multi-sheet Excel files by splitting them and cleaning each sheet individually.
"""

import sys
import os
import pandas as pd
from pathlib import Path
import shutil
import glob
import openpyxl
import logging
from crewai import Agent, Task, Crew, Process, LLM
from crewai_tools import MCPServerAdapter
from mcp import StdioServerParameters
from typing import List, Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    
)
logger = logging.getLogger(__name__)

def split_excel_by_sheets(input_file_path: str, temp_dir: str = "temp_sheets") -> List[Tuple[str, str, str]]:
    """
    Split an Excel file with multiple sheets into separate Excel files.
    
    Args:
        input_file_path (str): Path to the input Excel file
        temp_dir (str): Directory to save the split files
    
    Returns:
        List[Tuple[str, str, str]]: List of tuples containing (sheet_name, split_file_path, output_file_name)
    """
    # Create temp directory if it doesn't exist
    os.makedirs(temp_dir, exist_ok=True)
    
    # Get the base filename without extension
    base_filename = Path(input_file_path).stem
    
    split_info = []
    
    try:
        # Load the Excel file with openpyxl to check sheet visibility
        wb = openpyxl.load_workbook(input_file_path, read_only=True)
        excel_file = pd.ExcelFile(input_file_path)
        
        print(f"Found {len(excel_file.sheet_names)} sheets in {input_file_path}")
        
        for sheet_name in excel_file.sheet_names:
            # Skip all types of hidden sheets (hidden and veryHidden)
            if wb[sheet_name].sheet_state != 'visible':
                print(f"Skipping hidden sheet ({wb[sheet_name].sheet_state}): {sheet_name}")
                continue
            # Clean sheet name for filename (remove invalid characters)
            clean_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_', '+')).rstrip()
            clean_sheet_name = clean_sheet_name.replace(' ', '_')
            
            # Create split file path
            split_filename = f"{base_filename}_{clean_sheet_name}.xlsx"
            split_file_path = os.path.join(temp_dir, split_filename)
            
            # Create output file name (what the agent will produce)
            output_filename = f"cleaned_{base_filename}_{clean_sheet_name}.xlsx"
            
            # Read the specific sheet and save as new Excel file preserving all data
            df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=None, keep_default_na=False)
            
            # Log original data dimensions
            logger.info(f"Sheet '{sheet_name}': Original data dimensions {df.shape[0]} rows x {df.shape[1]} columns")
            
            with pd.ExcelWriter(split_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            # Validate split file was created correctly
            try:
                test_df = pd.read_excel(split_file_path, sheet_name=sheet_name, header=None, keep_default_na=False)
                if test_df.shape != df.shape:
                    logger.warning(f"Sheet '{sheet_name}': Dimension mismatch after split - Original: {df.shape}, Split: {test_df.shape}")
                else:
                    logger.info(f"Sheet '{sheet_name}': Split file validated successfully")
            except Exception as e:
                logger.error(f"Sheet '{sheet_name}': Failed to validate split file: {str(e)}")
            
            split_info.append((sheet_name, os.path.abspath(split_file_path), output_filename))
            print(f"Created: {split_filename} -> Output will be: {output_filename}")
    
    except Exception as e:
        print(f"Error splitting Excel file: {str(e)}")
        raise
    
    return split_info

def process_single_sheet_file(input_file: str, output_file: str, sheet_name: str) -> bool:
    """
    Process a single Excel file (containing one sheet) using the agent workflow.
    
    Args:
        input_file (str): Path to input Excel file
        output_file (str): Path to output Excel file
        sheet_name (str): Name of the original sheet (for context)
        
    Returns:
        bool: True if successful, False otherwise
    """
    
    # Convert to absolute paths
    input_file = os.path.abspath(input_file)
    output_file = os.path.abspath(output_file)
    
    # Validate input file exists
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return False
    
    # Validate input file is readable
    try:
        test_wb = openpyxl.load_workbook(input_file, read_only=True)
        test_wb.close()
    except Exception as e:
        print(f"❌ Cannot read input file {input_file}: {str(e)}")
        return False
    
    server_params = StdioServerParameters(
        command="python3",
        args=["-m", "excel_mcp", "stdio"],
        env={"UV_PYTHON": "3.12", **os.environ},
    )
    
    llm = LLM(
        model="bedrock/us.anthropic.claude-sonnet-4-20250514-v1:0",
        aws_region_name="us-east-1",
        temperature=0.2,
    )
    
    try:
        print(f"\n{'='*60}")
        print(f"Processing Sheet: {sheet_name}")
        print(f"Input File: {os.path.basename(input_file)}")
        print(f"Output File: {os.path.basename(output_file)}")
        print(f"{'='*60}")
        
        with MCPServerAdapter(server_params) as mcp_tools:
            print(f"Available tools: {[tool.name for tool in mcp_tools]}")

            # Agent 1: Excel Reader
            excel_reader = Agent(
                role='Excel File Reader',
                goal='Read Excel files and identify column headers for targeted data extraction',
                backstory=f"""You are an expert Excel file reader processing the '{sheet_name}' sheet.
                You specialize in reading Excel files efficiently by targeting specific data ranges.
                You avoid reading massive ranges that could cause performance issues. You focus on
                getting workbook metadata first, then reading smaller, targeted ranges to understand
                the data structure and identify column headers.
                You are particularly skilled at identifying column headers and understanding the full
                horizontal extent of tabular data to locate specific columns.""",
                tools=mcp_tools,
                verbose=True,
                allow_delegation=False,
                llm=llm
            )

            # Agent 2: Column Identifier and Data Processor
            data_processor = Agent(
                role='Column Identification Specialist',
                goal='Identify available columns from the target set: Group Name, Placement Name, SOV, Start Date, End Date, Units, Cost, and Rate',
                backstory=f"""You are a column identification expert processing the '{sheet_name}' sheet.
                You specialize in analyzing Excel column headers to identify up to 8 specific columns:
                1. Group Name (may also be called: Package Name, Group, Package, Campaign Group)
                2. Placement Name (may also be called: Placement, Ad Placement, Creative Name, Asset Name)
                3. SOV (may also be called: Share of Voice, SOV%, Share, Voice Share)
                4. Start Date (may also be called: Start, Begin Date, Campaign Start, Flight Start)
                5. End Date (may also be called: End, Finish Date, Campaign End, Flight End)
                6. Units (may also be called: Budget, Quantity, Volume, Impressions, Spots)
                7. Cost (may also be called: Price, Amount, Total Cost, Investment, Spend)
                8. Rate (may also be called: CPM, CPC, CPV, Unit Rate, Price Rate, Media Rate)
                
                IMPORTANT: Not all sheets will contain all 8 columns. You should identify and extract
                whichever columns are present from this target list. If only 3 or 4 of these columns
                exist in a sheet, that's perfectly acceptable - extract what's available.
                You are intelligent about recognizing these columns even with different naming conventions.
                You analyze header rows to map the actual column names to these target columns.
                You identify the exact column positions (like A, B, C, etc.) for each available target column.""",
                tools=mcp_tools,
                verbose=True,
                allow_delegation=False,
                llm=llm
            )

            # Agent 3: Selective Excel Writer
            excel_writer = Agent(
                role='Selective Excel File Writer',
                goal='Create new Excel files with ONLY the available identified columns and ALL their data',
                backstory=f"""You are a selective Excel file creation specialist processing the '{sheet_name}' sheet.
                You take whichever columns were identified from the target set (Group Name, Placement Name, SOV,
                Start Date, End Date, Units, Cost, Rate) and create new, properly formatted Excel files containing
                ONLY these available columns.
                
                IMPORTANT: Not all sheets will have all 8 target columns. You should work with whatever
                columns were successfully identified - whether that's 2, 4, 6, or all 8 columns.
                
                You are CRITICAL about data preservation for the identified columns - you MUST ensure that
                EVERY SINGLE ROW of data from the available columns is written to the output file.
                You preserve the original column headers and extract ALL rows of data for the identified columns only.
                You create a clean, focused output with just the available target columns and all their associated data.""",
                tools=mcp_tools,
                verbose=True,
                allow_delegation=False,
                llm=llm
            )

            # Task 1: Read Excel File Metadata and Identify Column Headers
            read_task = Task(
                description=f"""
                Read the Excel file located at '{input_file}' which contains data from the original '{sheet_name}' sheet.
                
                Your task is to:
                1. Use get_workbook_metadata tool to understand the file structure and sheet names
                2. For each sheet, read a wide sample range (e.g., A1:AC20) to capture all possible column headers
                3. Focus on identifying where the column headers are located and what they are called
                4. Look for headers that might correspond to these 8 target columns:
                   - Group Name (or Package Name, Group, Package, Campaign Group)
                   - Placement Name (or Placement, Ad Placement, Creative Name, Asset Name)
                   - SOV (or Share of Voice, SOV%, Share, Voice Share)
                   - Start Date (or Start, Begin Date, Campaign Start, Flight Start)
                   - End Date (or End, Finish Date, Campaign End, Flight End)
                   - Units (or Budget, Quantity, Volume, Impressions, Spots)
                   - Cost (or Price, Amount, Total Cost, Investment, Spend)
                   - Rate (or CPM, CPC, CPV, Unit Rate, Price Rate, Media Rate)
                5. Identify the row where actual data starts (after headers)
                6. Provide detailed information about all column headers found
                
                IMPORTANT: Read wide enough to capture all possible columns but focus on header identification.
                """,
                expected_output=f"Detailed metadata about the Excel file from '{sheet_name}' sheet, including all column headers found and their positions, with special attention to identifying the 8 target columns.",
                agent=excel_reader
            )

            # Task 2: Identify and Map the Available Target Columns
            process_task = Task(
                description=f"""
                Based on the column headers from the previous task, identify and map the available target columns from the '{sheet_name}' sheet.
                
                Your task is to:
                1. Analyze all the column headers to identify which ones correspond to these target columns:
                   - Group Name (variations: Package Name, Group, Package, Campaign Group)
                   - Placement Name (variations: Placement, Ad Placement, Creative Name, Asset Name)
                   - SOV (variations: Share of Voice, SOV%, Share, Voice Share)
                   - Start Date (variations: Start, Begin Date, Campaign Start, Flight Start)
                   - End Date (variations: End, Finish Date, Campaign End, Flight End)
                   - Units (variations: Budget, Quantity, Volume, Impressions, Spots)
                   - Cost (variations: Price, Amount, Total Cost, Investment, Spend)
                   - Rate (variations: CPM, CPC, CPV, Unit Rate, Price Rate, Media Rate)
                
                2. For each target column that IS FOUND in the sheet, identify:
                   - The exact column letter (A, B, C, etc.)
                   - The actual header name found in the sheet
                   - The row where the header is located
                   - The row where data starts
                
                3. Determine the data range for extraction (start row to end row)
                4. Create a mapping of available target columns to actual column positions
                5. Report which target columns were found and which ones are missing
                6. IMPORTANT: It's completely normal if only some of the target columns are present
                
                Focus on intelligent matching - be flexible with column name variations but ensure accuracy.
                Extract whatever target columns are available in this sheet.
                """,
                expected_output=f"A complete mapping of the available target columns to their actual positions in the '{sheet_name}' sheet, including column letters, header names, and data range information. Report which target columns were found and which ones are missing (missing columns are acceptable).",
                agent=data_processor
            )

            # Task 3: Extract and Write Only the Available Target Columns
            write_task = Task(
                description=f"""
                Create a new Excel file containing ONLY the available target columns and ALL their data from the '{sheet_name}' sheet.
                
                Your task is to:
                1. Use the column mapping from the previous task to identify the exact columns to extract
                2. For each target column that was found in the sheet:
                   - Read the complete column data using read_data_from_excel
                   - Include the header row and ALL data rows for that column
                3. Create a new workbook using create_workbook
                4. Write ONLY the identified target columns to the new file in this preferred order (skip missing ones):
                   - Group Name, Placement Name, SOV, Start Date, End Date, Units, Cost, Rate
                5. Preserve the original header names as found in the source
                6. Include ALL rows of data for these columns (no data should be skipped)
                7. Save the result to '{output_file}'
                8. IMPORTANT: Work with whatever columns were found - could be 2, 4, 6, or all 8 columns
                
                CRITICAL: Extract ALL data rows for the identified columns - every single row of data must be preserved.
                The output should be a clean, focused dataset with only the available target columns and all their data.
                
                The output file name '{os.path.basename(output_file)}' is specifically designed to map back to the original '{sheet_name}' sheet.
                """,
                expected_output=f"A new Excel file '{output_file}' containing ONLY the available target columns (from: Group Name, Placement Name, SOV, Start Date, End Date, Units, Cost, Rate) with ALL their data from the '{sheet_name}' sheet. The file should have clean headers and complete data for each identified column, regardless of how many target columns were found.",
                agent=excel_writer
            )

            # Create crew with all agents and tasks
            crew = Crew(
                agents=[excel_reader, data_processor, excel_writer],
                tasks=[read_task, process_task, write_task],
                verbose=True
            )

            # Execute the crew
            result = crew.kickoff()
            
            # Validate output file was created and contains data
            if os.path.exists(output_file):
                try:
                    # Quick validation of output file
                    output_wb = openpyxl.load_workbook(output_file, read_only=True)
                    output_ws = output_wb.active
                    
                    # Enhanced validation for selective column extraction
                    print(f"✅ Output dimensions: {output_ws.max_row} rows x {output_ws.max_column} columns")
                    
                    # Validate that we have the expected number of columns (up to 8)
                    if output_ws.max_column > 8:
                        print(f"⚠️  Warning: Output has more than 8 columns ({output_ws.max_column}) - expected maximum 8")
                    elif output_ws.max_column < 1:
                        print(f"⚠️  Warning: Output has no columns")
                    else:
                        print(f"✅ Column count validation: {output_ws.max_column} columns (expected up to 8)")
                    
                    # Count data rows (excluding header)
                    data_rows = max(0, output_ws.max_row - 1) if output_ws.max_row > 1 else 0
                    print(f"✅ Data rows: {data_rows} (plus 1 header row)")
                    
                    # Check if we have headers
                    if output_ws.max_row >= 1:
                        header_row = []
                        for col in range(1, output_ws.max_column + 1):
                            cell_value = output_ws.cell(row=1, column=col).value
                            header_row.append(str(cell_value) if cell_value else "")
                        print(f"✅ Headers found: {', '.join(header_row)}")
                    
                    if output_ws.max_row < 2:
                        print(f"⚠️  Warning: Output has very few rows - may indicate no data extracted")
                    
                    output_wb.close()
                except Exception as e:
                    print(f"⚠️  Warning: Could not validate output file: {str(e)}")
            else:
                print(f"❌ Output file was not created: {output_file}")
                return False
            
            print(f"✅ Successfully processed sheet '{sheet_name}'")
            return True
            
    except Exception as e:
        print(f"❌ Error processing sheet '{sheet_name}': {str(e)}")
        return False

def process_multi_sheet_excel(input_excel_path: str, output_directory: str = "processed_sheets"):
    """
    Main function to process a multi-sheet Excel file by splitting it and processing each sheet individually.
    
    Args:
        input_excel_path (str): Path to the input Excel file with multiple sheets
        output_directory (str): Directory to save processed files
    """
    print(f"🚀 Starting multi-sheet Excel processing...")
    print(f"Input file: {input_excel_path}")
    
    # Check if input file exists
    if not os.path.exists(input_excel_path):
        print(f"❌ Input file not found: {input_excel_path}")
        return
    
    # Create output directory
    os.makedirs(output_directory, exist_ok=True)
    
    # Create temporary directory for split sheets
    temp_dir = "temp_split_sheets"
    
    successful_files = []
    failed_files = []
    
    try:
        # Step 1: Split sheets into individual files
        print(f"\n✂️ Splitting sheets into individual files...")
        split_info = split_excel_by_sheets(input_excel_path, temp_dir)
        
        # Step 2: Process each file individually
        print(f"\n🔄 Processing {len(split_info)} sheets sequentially...")
        
        for i, (sheet_name, split_file_path, output_filename) in enumerate(split_info, 1):
            # Create full output path
            output_file_path = os.path.join(output_directory, output_filename)
            
            print(f"\n[{i}/{len(split_info)}] Processing '{sheet_name}' sheet...")
            
            success = process_single_sheet_file(split_file_path, output_file_path, sheet_name)
            
            if success:
                successful_files.append((sheet_name, output_filename))
                print(f"✅ Sheet '{sheet_name}' completed successfully -> {output_filename}")
            else:
                failed_files.append((sheet_name, output_filename))
                print(f"❌ Sheet '{sheet_name}' failed")
    
    finally:
        # Clean up temporary directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n🧹 Cleaned up temporary files")
    
    # Summary
    print(f"\n{'='*60}")
    print(f"📈 PROCESSING SUMMARY")
    print(f"{'='*60}")
    print(f"✅ Successfully processed: {len(successful_files)} sheets")
    print(f"❌ Failed: {len(failed_files)} sheets")
    
    if successful_files:
        print(f"\n✅ Successful sheets:")
        for sheet_name, output_file in successful_files:
            print(f"  - '{sheet_name}' -> {output_file}")
    
    if failed_files:
        print(f"\n❌ Failed sheets:")
        for sheet_name, output_file in failed_files:
            print(f"  - '{sheet_name}' -> {output_file}")
    
    print(f"\n📁 Output directory: {os.path.abspath(output_directory)}")
    print(f"🎉 Multi-sheet processing completed!")

# Configuration section
CONFIG = {
    'input_dir': 'input_excels',  # Directory containing Excel files
    'output_base': 'processed_files'  # Base output directory
}

def find_excel_files(path):
    """Find all Excel files in a directory or return single file path."""
    if os.path.isfile(path):
        return [path] if path.lower().endswith(('.xlsx', '.xls')) else []
    return glob.glob(os.path.join(path, '*.xlsx')) + glob.glob(os.path.join(path, '*.xls'))

def main():
    """Main function to process all Excel files in input directory."""
    
    print("📊 Batch Excel Processor")
    print("=" * 35)
    
    # Find all Excel files
    excel_files = find_excel_files(CONFIG['input_dir'])
    
    if not excel_files:
        print(f"❌ No Excel files found in: {CONFIG['input_dir']}")
        return
    
    print(f"\nFound {len(excel_files)} Excel files to process:")
    for file in excel_files:
        print(f"  - {os.path.basename(file)}")
    
    # Create base output directory
    os.makedirs(CONFIG['output_base'], exist_ok=True)
    
    # Process each file
    for input_file in excel_files:
        try:
            base_name = Path(input_file).stem
            output_dir = os.path.join(CONFIG['output_base'], base_name)
            
            print(f"\n{'='*60}")
            print(f"Processing: {os.path.basename(input_file)}")
            print(f"Output to: {output_dir}/")
            print(f"{'='*60}")
            
            process_multi_sheet_excel(input_file, output_dir)
        except Exception as e:
            print(f"❌ Error processing {os.path.basename(input_file)}: {str(e)}")
            continue

if __name__ == "__main__":
    main()
